"""Write a multi-sheet spreadsheet, with a CSV fallback.

``openpyxl`` is not part of Slicer's Python, so it is installed on demand the
same way the rest of this project's optional dependencies are. When that is not
possible -- offline, or a locked-down install -- the same tables are written as
one CSV per sheet rather than the run producing nothing: the numbers are the
deliverable, the .xlsx is only the container.
"""

from __future__ import annotations

import csv
import io
import logging
import os
from typing import Any, Dict, List, Optional, Sequence, Tuple

logger = logging.getLogger(__name__)

#: A sheet is (title, blocks); a block is (caption or "", column names, rows).
#: Several blocks per sheet so one tab can carry a table and its aggregates --
#: "one sub-tab for BIC" without hiding the per-side totals somewhere else.
Block = Tuple[str, Sequence[str], Sequence[Dict[str, Any]]]
Sheet = Tuple[str, Sequence[Block]]

#: Column width, in characters, at which a cell is treated as prose: wrapped
#: rather than allowed to set the column's width. Definition blocks are mostly
#: sentences, and one of them would otherwise stretch a column across the screen.
PROSE_WIDTH = 70


def _ensure_openpyxl():
    try:
        import openpyxl                                      # noqa: PLC0415
        return openpyxl
    except ImportError:
        pass
    try:
        import slicer                                        # noqa: PLC0415
        slicer.util.pip_install("openpyxl")
        import openpyxl                                      # noqa: PLC0415
        return openpyxl
    except Exception:
        logger.info("openpyxl unavailable; falling back to CSV", exc_info=True)
        return None


def _cell(value: Any) -> Any:
    if isinstance(value, bool):
        return "yes" if value else "NO"
    return "" if value is None else value


def write_workbook(path: str, sheets: Sequence[Sheet]) -> Tuple[str, List[str]]:
    """Write ``sheets`` to ``path``. Returns ``(written_path, notes)``."""
    notes: List[str] = []
    openpyxl = _ensure_openpyxl()
    if openpyxl is None:
        written = _write_csv_fallback(path, sheets)
        notes.append("openpyxl is not available, so the tables were written as "
                     "CSV files instead of one .xlsx.")
        return written, notes

    from openpyxl.styles import Alignment, Font               # noqa: PLC0415

    book = openpyxl.Workbook()
    book.remove(book.active)
    for title, blocks in sheets:
        sheet = book.create_sheet(title[:31])                 # Excel's limit
        widths: Dict[int, int] = {}
        row_index = 1
        for caption, columns, rows in blocks:
            if caption:
                cell = sheet.cell(row=row_index, column=1, value=caption)
                cell.font = Font(bold=True, size=12)
                row_index += 1
            for column_index, name in enumerate(columns, start=1):
                cell = sheet.cell(row=row_index, column=column_index, value=name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                widths[column_index] = max(widths.get(column_index, 0), len(str(name)))
            row_index += 1
            for row in rows:
                for column_index, name in enumerate(columns, start=1):
                    value = _cell(row.get(name))
                    cell = sheet.cell(row=row_index, column=column_index, value=value)
                    text = str(value)
                    # Prose (a definition, a failure reason) is wrapped and
                    # top-aligned, so it is readable in the cell instead of
                    # running under the next column or being clipped. Numbers
                    # and short labels are left alone -- wrapping those would
                    # only make the rows taller.
                    if len(text) > PROSE_WIDTH:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        widths[column_index] = max(widths.get(column_index, 0),
                                                   PROSE_WIDTH)
                    else:
                        widths[column_index] = max(widths.get(column_index, 0),
                                                   len(text))
                row_index += 1
            row_index += 1                                    # blank separator
        for column_index, width in widths.items():
            sheet.column_dimensions[
                openpyxl.utils.get_column_letter(column_index)
            ].width = min(max(width + 2, 9), 60)
        sheet.freeze_panes = "A2"
    try:
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        book.save(path)
    except PermissionError:
        # Almost always the file open in Excel, which locks it. Say so, and say
        # what to do, rather than surfacing a bare OS error.
        raise PermissionError(
            f"Cannot write {os.path.basename(path)} -- it is open in another "
            f"program. Close it and run the analysis again.")
    return path, notes


def _write_csv_fallback(path: str, sheets: Sequence[Sheet]) -> str:
    base, _unused = os.path.splitext(path)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    written: List[str] = []
    for title, blocks in sheets:
        target = f"{base}_{title}.csv"
        with io.open(target, "w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            for caption, columns, rows in blocks:
                if caption:
                    writer.writerow([caption])
                writer.writerow(list(columns))
                for row in rows:
                    writer.writerow([_cell(row.get(name)) for name in columns])
                writer.writerow([])
        written.append(target)
    return written[0] if written else path
